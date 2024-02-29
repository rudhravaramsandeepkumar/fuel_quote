import gc
from flask import Blueprint, render_template, request, redirect, url_for, session, flash, send_file, make_response, \
    Response, jsonify
from sqlalchemy.orm.exc import NoResultFound, MultipleResultsFound

from .models.UserCredentials import UserCredentials
from .models.ClientInformation import ClientInformation
from .models.price_management import PriceManagement
from .models.FuelQuote import FuelQuote
from .models.Otp_Manager import OtpManager
from . import LOG, APP
from .models import db
from datetime import datetime
from .helpers import get_random_numbers, validate_gmail
from .security import Encryption
from .otp_generator import generate_otp, send_email_fun
import datetime as td
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
import openpyxl

bp = Blueprint('view', __name__, url_prefix='/fuel_quote', template_folder="./templates", static_folder="./static")


@bp.route('/check', methods=["GET", "POST"])
def enc_key():
    data = request.json
    print(data)
    reg_username = data['reg_username']
    reg_email = data['reg_email']
    print(data)
    try:
        users = UserCredentials.query.filter_by(username=reg_username).one()
        print(users.username)
        msg = ''
        if users is not None:
            msg = "user_name"
            return jsonify({"msg": msg})
    except NoResultFound:
        pass
    except Exception as e:
        LOG.error("Error occurred while login ")
        LOG.error(e, exc_info=True)
    finally:
        db.session.close()
    try:
        users2 = UserCredentials.query.filter_by(email_id=reg_email).one()
        if users2 is not None:
            msg = "email_id"
            return jsonify({
                "msg": msg})
        else:
            msg = "clear"
            return jsonify({
                "msg": msg})
    except NoResultFound:
        pass
    except Exception as e:
        LOG.error("Error occurred while login ")
        LOG.error(e, exc_info=True)
    finally:
        db.session.close()
    Status = {"msg": "msg"}
    return jsonify(Status)


@bp.route('/', methods=["GET", "POST"])
def login():
    if request.method == "GET":
        invalid_msg = "hidden"
        access_name = "hidden"
        access_pass = "hidden"
        return render_template("login.html", invalid_msg=invalid_msg, access_name=access_name, access_pass=access_pass)
    if request.method == "POST":
        username = request.form.get('username')
        password = request.form.get('password')
        if username == '':
            invalid_msg = "hidden"
            access_name = ""
            access_pass = "hidden"
            return render_template("login.html", invalid_msg=invalid_msg, access_name=access_name,
                                   access_pass=access_pass)
        if password == '':
            invalid_msg = "hidden"
            access_name = "hidden"
            access_pass = ""
            return render_template("login.html", invalid_msg=invalid_msg, access_name=access_name,
                                   access_pass=access_pass)
        try:
            encrypted_pwd = Encryption().encrypt(password)
            users = UserCredentials.query.filter_by(username=username, password=encrypted_pwd).one()
            user_id_get = users.user_uid
            print('user_id_get', user_id_get)
            if users is not None:
                if users.status == 'completed' and users.attributes_1 == 'user':
                    session["user"] = username
                    session["role"] = users.attributes_1
                    session["user_uid"] = users.user_uid
                    session["id"] = users.Id
                    return redirect(url_for("view.fuel_quote_history", user_uid=users.user_uid))
                if users.attributes_1 == 'admin':
                    print('kakakakak')
                    session["user"] = username
                    session["role"] = users.attributes_1
                    session["user_uid"] = users.user_uid
                    session["id"] = users.Id
                    session["display_name"] = users.username
                    return redirect(url_for("admin_view.admin_home", user_uid=users.user_uid))
                else:
                    users = ClientInformation.query.filter_by(user_id=user_id_get).one()
                    session["user"] = username
                    session["user_uid"] = users.user_id
                    session["username"] = users.username
                    session["email_id"] = users.email_id
                    session["id"] = users.Id
                    print('users.Id', users.Id)
                    UID = users.user_id
                    Username = users.username
                    Email_Id = users.email_id
                    First_Name = ""
                    Address_1 = ""
                    Address_2 = ""
                    country = "0"
                    state_code = ""
                    City_name = ""
                    Zip_Code = ""
                    Phone_Number = ""
                    return render_template("client_profile_management.html", UID=UID, Username=Username,
                                           Email_Id=Email_Id, First_Name=First_Name, Address_1=Address_1,
                                           Address_2=Address_2
                                           , country=country, state_code=state_code, City_name=City_name,
                                           Zip_Code=Zip_Code, Phone_Number=Phone_Number
                                           )
        except NoResultFound:
            pass
        except Exception as e:
            LOG.error("Error occurred while login ")
            LOG.error(e, exc_info=True)
        finally:
            db.session.close()
        invalid_msg = ""
        access_name = "hidden"
        access_pass = "hidden"
        return render_template("login.html", invalid_msg=invalid_msg, access_name=access_name, access_pass=access_pass)


@bp.route('/client_profile_management/<string:user_uid>', methods=['GET', 'POST'])
def client_profile_management(user_uid):
    print('user_uid', user_uid)
    if request.method == "GET":
        users = ClientInformation.query.filter_by(user_id=user_uid).one()
        if users is not None:
            if users.status == 'completed':
                session["user"] = users.username
                session["user_uid"] = users.user_id
                session["username"] = users.username
                session["email_id"] = users.email_id
                session["id"] = users.Id
                UID = users.user_id
                Username = users.username
                Email_Id = users.email_id
                OFIICE_NAME = users.office_name
                Address_1 = users.address_1
                Address_2 = users.address_2
                country = users.country
                state_code = users.state
                City_name = users.attributes_1
                Zip_Code = users.zipcode
                Phone_Number = users.mobile_number
                return render_template("client_profile_management.html", UID=UID, Username=Username,
                                       Email_Id=Email_Id, First_Name=OFIICE_NAME,
                                       Address_1=Address_1, Address_2=Address_2
                                       , country=country, state_code=state_code, City_name=City_name, Zip_Code=Zip_Code,
                                       Phone_Number=Phone_Number
                                       )
            else:
                session["user"] = users.username
                session["user_uid"] = users.user_id
                session["username"] = users.username
                session["email_id"] = users.email_id
                session["id"] = users.Id
                print('users.Id', users.Id)
                UID = users.user_id
                Username = users.username
                Email_Id = users.email_id
                OFIICE_NAME = ""
                Address_1 = ""
                Address_2 = ""
                country = "0"
                state_code = ""
                City_name = ""
                Zip_Code = ""
                Phone_Number = ""
                return render_template("client_profile_management.html", UID=UID, Username=Username,
                                       Email_Id=Email_Id, First_Name=OFIICE_NAME,
                                       Address_1=Address_1, Address_2=Address_2
                                       , country=country, state_code=state_code, City_name=City_name, Zip_Code=Zip_Code,
                                       Phone_Number=Phone_Number
                                       )


@bp.route('/client_registration', methods=['POST'])
def client_registration():
    if request.method == "POST":
        data = request.json
        print('data', data)
        if not (len(data['reg_username']) <= 50):
            msg = "User name is too long"
            Status = {"status": msg}
            return jsonify(Status)
        if not validate_gmail(data['reg_email']):
            msg = "Invalid Gmail"
            Status = {"status": msg}
            return jsonify(Status)

        header_id = datetime.now().strftime("%d%m%Y%H%M%S") + get_random_numbers(5)
        encrypted_pwd = Encryption().encrypt(data['reg_password'])
        head = UserCredentials(user_uid=header_id,
                               password=encrypted_pwd,
                               username=data['reg_username'],
                               email_id=data['reg_email'],
                               active='y',
                               status='pending',
                               attributes_1='user')
        try:
            db.session.add(head)
            db.session.commit()
            db.session.close()
        except Exception as e:
            LOG.error(e, exc_info=True)
            db.session.rollback()
            LOG.error("Error while pushing data")
            msg = "Failed"
            Status = {"status": msg}
            return jsonify(Status)
        head2 = ClientInformation(user_id=header_id,
                                  username=data['reg_username'],
                                  email_id=data['reg_email'],
                                  mobile_number="None",
                                  office_name="None",
                                  active='y',
                                  status='pending',
                                  zipcode="None",
                                  address_1="None",
                                  address_2="None",
                                  state="None",
                                  country="None")
        try:
            db.session.add(head2)
            db.session.commit()
            db.session.close()
            msg = "successfully registered"
            Status = {"status": msg}
            print('Status', Status)
            return jsonify(Status)
        except Exception as e:
            LOG.error(e, exc_info=True)
            db.session.rollback()
            msg = "Failed"
            Status = {"status": msg}
            print('Status', Status)
            return jsonify(Status)


@bp.route('/Update_client/<string:user_uid>', methods=['POST'])
def update_client(user_uid):
    if request.method == "POST":
        data = request.json
        First_Name = data["First_Name"]
        Address_1 = data["Address_1"]
        Address_2 = data["Address_2"]
        Zip_Code = data["Zip_Code"]
        City_name = data["City_name"]
        if not (len(First_Name) <= 50):
            msg = "First name is too long"
            Status = {"status": msg}
            return jsonify(Status)
        if not (len(Address_1) <= 100):
            msg = "Address 1 name is too long"
            Status = {"status": msg}
            return jsonify(Status)
        if not (len(Address_2) <= 100):
            msg = "Address 2 is too long"
            Status = {"status": msg}
            return jsonify(Status)
        if not (len(Zip_Code) <= 9):
            msg = "Zip name is too long"
            Status = {"status": msg}
            return jsonify(Status)
        if not (len(City_name) <= 100):
            msg = "City name is too long"
            Status = {"status": msg}
            return jsonify(Status)
        users = ClientInformation.query.filter_by(user_id=user_uid).one()
        head = ClientInformation(Id=users.Id,
                                 username=users.username,
                                 mobile_number=data['Phone_Number'],
                                 office_name=data['First_Name'],
                                 active='y',
                                 status='completed',
                                 zipcode=data['Zip_Code'],
                                 address_1=data['Address_1'],
                                 address_2=data['Address_2'],
                                 state=data['state_code'],
                                 country=data['country'],
                                 attributes_1=data['City_name'])
        try:
            db.session.merge(head)
            db.session.commit()
            db.session.close()
        except Exception as e:
            LOG.error(e, exc_info=True)
            db.session.rollback()
            LOG.error("Error while pushing data")
            msg = "Failed"
            Status = {"status": msg}
            return jsonify(Status)
        users = UserCredentials.query.filter_by(user_uid=user_uid).one()
        head1 = UserCredentials(Id=users.Id,
                                username=users.username,
                                active='y',
                                status='completed')
        try:
            db.session.merge(head1)
            db.session.commit()
            db.session.close()
            msg = "successfully registered"
            Status = {"status": msg}
            return jsonify(Status)
        except Exception as e:
            LOG.error(e, exc_info=True)
            db.session.rollback()
            LOG.error("Error while pushing data")
            msg = "Failed"
            Status = {"status": msg}
            return jsonify(Status)


@bp.route('/fuel_quote_history/<string:user_uid>', methods=['GET', 'POST'])
def fuel_quote_history(user_uid):
    if request.method == "GET":
        fq = FuelQuote.query.filter_by(user_id=user_uid).all()
        users = ClientInformation.query.filter_by(user_id=user_uid).one()
        main_table = []
        inner_loop = []
        first_username = users.office_name
        for Quote in fq:
            print(Quote)
            inner_loop.append(Quote.fuel_quote_id)
            inner_loop.append(Quote.fuel_type)
            inner_loop.append(Quote.cur_gallon_price)
            inner_loop.append(Quote.quantity)
            inner_loop.append(Quote.total_price)
            inner_loop.append(Quote.created_date)
            inner_loop.append(Quote.delivery_date)
            main_table.append(inner_loop)
            inner_loop = []
        count_row = len(main_table)
        return render_template("fuel_quote_history.html", UID=user_uid, main_table=main_table, count_row=count_row,
                               first_username=first_username)


@bp.route('/fuel_quote_form/<string:user_uid>', methods=['GET', 'POST'])
def fuel_quote_form(user_uid):
    try:
        users = ClientInformation.query.filter_by(user_id=user_uid).one()
        id = users.user_id
        Office_Name = users.office_name
        delivery_address = users.address_1
        Country = users.country
        state = users.state
        zipcode = users.zipcode
        city = users.attributes_1
        email_id = users.email_id
        mobile_number = users.mobile_number
        quantity_quote = ''
        users = PriceManagement.query.filter_by(fuel_type="diesel").one()
        current_price_ = users.price_
        available_quantity = users.quantity_
        if request.method == "GET":
            return render_template("fuel_quote_form.html", UID=id, id=id, Office_Name=Office_Name,
                                   delivery_address=delivery_address,
                                   Country=Country, state=state, zipcode=zipcode, city=city, email_id=email_id,
                                   mobile_number=mobile_number, quantity_quote=quantity_quote, edit_="",
                                   current_price_=current_price_, available_quantity=available_quantity)
    except NoResultFound:
        pass
    except Exception as e:
        LOG.error("Error occurred while login ")
        LOG.error(e, exc_info=True)
    finally:
        db.session.close()
    if request.method == "POST":
        data = request.json
        fuel_quote_id = "BKID" + datetime.now().strftime("%d%m%Y%H%M%S") + get_random_numbers(2)
        print(data['bdate'])

        delivery_date = datetime.strptime(data['bdate'], '%Y-%m-%d')
        # dnd = datetime.strptime(delivery_date, '%Y-%m-%d %H:%M:%S.%f')
        final_end_date = delivery_date.strftime("%Y-%m-%d %H:%M:%S.%f")
        dnd = datetime.strptime(final_end_date, '%Y-%m-%d %H:%M:%S.%f')
        print('final_end_date', final_end_date)
        users = PriceManagement.query.filter_by(fuel_type="diesel").one()
        head = PriceManagement(Id=users.Id,
                               quantity_=data['available_quantity'])
        try:
            db.session.merge(head)
            db.session.commit()
            db.session.close()
        except Exception as e:
            LOG.error(e, exc_info=True)
            db.session.rollback()
        head = FuelQuote(fuel_quote_id=fuel_quote_id,
                         user_id=id,
                         quantity=data['quantity_quote'],
                         cur_gallon_price=data['Gallon_price'],
                         total_price=data['Total_Amount_Due'],
                         state=state,
                         country=Country,
                         status='completed',
                         address=delivery_address,
                         phone_number=mobile_number,
                         fuel_type=data['fuel_type'],
                         cur_gst=data['cur_gst'],
                         delivery_date=dnd)
        try:
            db.session.merge(head)
            db.session.commit()
            db.session.close()
            msg = "successfully Saved"
            Status = {"status": msg}
            return jsonify(Status)
        except Exception as e:
            LOG.error(e, exc_info=True)
            db.session.rollback()
            LOG.error("Error while pushing data")
            msg = "Failed"
            Status = {"status": msg}
            return jsonify(Status)


@bp.route('/repeat_order/<string:booking_id>', methods=['GET', 'POST'])
def repeat_order(booking_id):
    try:
        client_info = FuelQuote.query.filter_by(fuel_quote_id=booking_id).one()
        users = ClientInformation.query.filter_by(user_id=client_info.user_id).one()
        id = users.user_id
        Office_Name = users.office_name
        delivery_address = users.address_1
        Country = users.country
        state = users.state
        zipcode = users.zipcode
        city = users.attributes_1
        email_id = users.email_id
        mobile_number = users.mobile_number
        quantity_quote = client_info.quantity
        users = PriceManagement.query.filter_by(fuel_type="diesel").one()
        current_price_ = users.price_
        available_quantity = users.quantity_
        if request.method == "GET":
            return render_template("fuel_quote_form.html", UID=id, id=id, Office_Name=Office_Name,
                                   delivery_address=delivery_address,
                                   Country=Country, state=state, zipcode=zipcode, city=city, email_id=email_id,
                                   mobile_number=mobile_number, quantity_quote=quantity_quote, edit_="disabled",
                                   current_price_=current_price_,
                                   available_quantity=available_quantity)
    except NoResultFound:
        pass
    except Exception as e:
        LOG.error("Error occurred while login ")
        LOG.error(e, exc_info=True)
    finally:
        db.session.close()


@bp.route('/Get_Quote', methods=['POST'])
def get_quote():
    if request.method == "POST":
        data = request.json
        print(data)
        fq = FuelQuote.query.filter_by(user_id=data["user_uid"]).all()
        users = PriceManagement.query.filter_by(fuel_type="diesel").one()
        available_quantity = int(users.quantity_) - int(data['quantity_quote'])

        current_price_per_gallon = float(users.price_)
        Company_Profit_Factor = 0.1
        if data['state_code'] == 'TX':
            location_factor = 0.02
        else:
            location_factor = 0.04
        if fq:
            rate_history_factor = 0.01
        else:
            rate_history_factor = 0
        if int(data['quantity_quote']) > 1000:
            gallons_requested_factor = 0.02
        else:
            gallons_requested_factor = 0.03
        margin_ = current_price_per_gallon * (
                    location_factor - rate_history_factor + gallons_requested_factor + Company_Profit_Factor)
        Suggested_Price = current_price_per_gallon + margin_
        Total_Amount_Due = int(data['quantity_quote']) * Suggested_Price
        final_ = {"margin_": round(margin_, 2), "Suggested_Price": Suggested_Price,
                  "Total_Amount_Due": Total_Amount_Due, "available_quantity": available_quantity}
        return jsonify(final_)


@bp.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == "GET":
        return render_template("forgot_password.html")
    if request.method == "POST":
        data = request.json
        reg_gmail = data['user_name']
        otp_ = data['otp']
        users2 = UserCredentials.query.filter_by(email_id=reg_gmail).one()
        id_ = users2.user_uid
        try:
            otp_manager = OtpManager.query.filter_by(otp=otp_).one()
            otp_id_ = otp_manager.ott_id
            fmt = '%Y-%m-%d %H:%M:%S.%f'
            d1 = datetime.strptime(str(otp_manager.created_date), fmt)
            d2 = datetime.strptime(str(datetime.now()), fmt)
            diff = d2 - d1
            diff_minutes = (diff.days * 24 * 60) + (diff.seconds / 60)
            if otp_manager is not None:
                if (str(otp_manager.fuel_quote_id).strip() == str(id_).strip()) and not diff_minutes > 5:
                    if otp_manager.status == "active":
                        encrypted_pwd = Encryption().encrypt(data['pass_word'])
                        head = UserCredentials(Id=users2.Id,
                                               password=encrypted_pwd)
                        try:
                            db.session.merge(head)
                            db.session.commit()
                            db.session.close()
                            print('here')
                        except Exception as e:
                            LOG.error(e, exc_info=True)
                            db.session.rollback()
                            LOG.error("Error while pushing data")
                            msg = "Error while pushing data"
                            Status = {"msg": msg}
                            return jsonify(Status)
                        head = OtpManager(ott_id=otp_id_,
                                          status='Inactive')
                        try:
                            db.session.merge(head)
                            db.session.commit()
                            db.session.close()
                        except Exception as e:
                            LOG.error(e, exc_info=True)
                            db.session.rollback()
                            LOG.error("Error while pushing data")
                            msg = "Error while pushing data"
                            Status = {"msg": msg}
                            return jsonify(Status)
                        msg = "Successfully Changed"
                        return jsonify({"msg": msg})
                    else:
                        msg = "not_valid"
                        return jsonify({"msg": msg})
                else:
                    msg = "Expired"
                    return jsonify({"msg": msg})
            else:
                msg = "not_found"
                return jsonify({"msg": msg})

        except NoResultFound:
            msg = "not_found"
            return jsonify({
                "msg": msg})
            pass
        except Exception as e:
            LOG.error("Error occurred while login ")
            LOG.error(e, exc_info=True)
        finally:
            db.session.close()


@bp.route('/send_otp', methods=["GET", "POST"])
def send_otp():
    data = request.json
    reg_email = data['reg_email']
    try:
        users2 = UserCredentials.query.filter_by(email_id=reg_email).one()
        if users2 is not None:
            id_ = users2.user_uid
            name_ = users2.username
            msg = "clear"
            otp_ = generate_otp(id_)
            send_email_fun(APP.config['FROM_MAIL'], reg_email, name_, otp_, datetime.now(), '5 Min')
            return jsonify({"msg": msg})
        else:
            msg = "email_id"
            return jsonify({
                "msg": msg})
    except NoResultFound:
        msg = "email_id"
        return jsonify({
            "msg": msg})
        pass
    except Exception as e:
        LOG.error("Error occurred while login ")
        LOG.error(e, exc_info=True)
    finally:
        db.session.close()
    Status = {"msg": "msg"}
    return jsonify(Status)


@bp.route('/download_report/<string:user_uid>', methods=["GET", "POST"])
def download_report(user_uid):
    fuel_quote_id, user_id, quantity, cur_gallon_price, total_price, state, country, address, \
        phone_number, fuel_type, cur_gst, created_date, delivery_date = [], [], [], [], [], [], [], [], [], [], [], [], []

    claims = FuelQuote.query.filter_by(user_id=user_uid).all()

    for claim in claims:
        fuel_quote_id.append(claim.fuel_quote_id)
        user_id.append(claim.user_id)
        quantity.append(claim.quantity)
        cur_gallon_price.append(claim.cur_gallon_price)
        total_price.append(claim.total_price)
        state.append(claim.state)
        country.append(claim.country)
        address.append(claim.address)
        phone_number.append(claim.phone_number)
        fuel_type.append(claim.fuel_type)
        cur_gst.append(claim.cur_gst)
        created_date.append(claim.created_date)
        delivery_date.append(claim.delivery_date)

    today = td.datetime.today()
    created_date_new = str(today.strftime("%d/%m/%Y %H:%M:%S")).replace(" ", "_").replace(":", "_").replace("/",
                                                                                                            "_")
    filepath = APP.config['ROOT_DIR'] + '/' + "report/fuelQuote_Report_" + str(
        created_date_new) + ".xlsx"
    wb = openpyxl.Workbook()
    sheet_obj = wb.active
    my_font = Font(name='Arial',
                   size=11,
                   bold=True,
                   italic=False,
                   vertAlign=None,
                   underline='none',
                   strike=False,
                   color='FFFFFF')
    my_font_1 = Font(name='Arial',
                     size=10,
                     bold=False,
                     italic=False,
                     vertAlign=None,
                     underline='none',
                     strike=False,
                     color='000000')
    my_pattern = PatternFill(start_color="2a6099", end_color="2a6099", fill_type="solid")
    my_pattern_warning = PatternFill(start_color="ffdbb6", end_color="ffdbb6", fill_type="solid")
    row_start = 1
    sheet_obj.cell(row=1, column=1).value = "Booking id"
    sheet_obj.cell(row=1, column=1).font = my_font
    sheet_obj.cell(row=1, column=1).fill = my_pattern
    sheet_obj.cell(row=1, column=2).value = "Ref Number"
    sheet_obj.cell(row=1, column=2).font = my_font
    sheet_obj.cell(row=1, column=2).fill = my_pattern
    sheet_obj.cell(row=1, column=3).value = "Quantity"
    sheet_obj.cell(row=1, column=3).font = my_font
    sheet_obj.cell(row=1, column=3).fill = my_pattern
    sheet_obj.cell(row=1, column=4).value = "Current Gallon Price"
    sheet_obj.cell(row=1, column=4).font = my_font
    sheet_obj.cell(row=1, column=4).fill = my_pattern
    sheet_obj.cell(row=1, column=5).value = "Total Price"
    sheet_obj.cell(row=1, column=5).font = my_font
    sheet_obj.cell(row=1, column=5).fill = my_pattern
    sheet_obj.cell(row=1, column=6).value = "State"
    sheet_obj.cell(row=1, column=6).font = my_font
    sheet_obj.cell(row=1, column=6).fill = my_pattern
    sheet_obj.cell(row=1, column=7).value = "Country"
    sheet_obj.cell(row=1, column=7).font = my_font
    sheet_obj.cell(row=1, column=7).fill = my_pattern
    sheet_obj.cell(row=1, column=8).value = "Address"
    sheet_obj.cell(row=1, column=8).font = my_font
    sheet_obj.cell(row=1, column=8).fill = my_pattern
    sheet_obj.cell(row=1, column=9).value = "Phone_number"
    sheet_obj.cell(row=1, column=9).font = my_font
    sheet_obj.cell(row=1, column=9).fill = my_pattern
    sheet_obj.cell(row=1, column=10).value = "Fuel Type"
    sheet_obj.cell(row=1, column=10).font = my_font
    sheet_obj.cell(row=1, column=10).fill = my_pattern
    sheet_obj.cell(row=1, column=11).value = "Current Gst"
    sheet_obj.cell(row=1, column=11).font = my_font
    sheet_obj.cell(row=1, column=11).fill = my_pattern
    sheet_obj.cell(row=1, column=12).value = "Created Date"
    sheet_obj.cell(row=1, column=12).font = my_font
    sheet_obj.cell(row=1, column=12).fill = my_pattern
    sheet_obj.cell(row=1, column=13).value = "Delivery Date"
    sheet_obj.cell(row=1, column=13).font = my_font
    sheet_obj.cell(row=1, column=13).fill = my_pattern

    for i in range(len(fuel_quote_id)):
        sheet_obj.cell(row=row_start + 1, column=1).value = fuel_quote_id[i]
        sheet_obj.cell(row=row_start + 1, column=1).font = my_font_1
        sheet_obj.cell(row=row_start + 1, column=2).value = user_id[i]
        sheet_obj.cell(row=row_start + 1, column=2).font = my_font_1
        sheet_obj.cell(row=row_start + 1, column=3).value = quantity[i]
        sheet_obj.cell(row=row_start + 1, column=3).font = my_font_1
        sheet_obj.cell(row=row_start + 1, column=4).value = cur_gallon_price[i]
        sheet_obj.cell(row=row_start + 1, column=4).font = my_font_1
        sheet_obj.cell(row=row_start + 1, column=5).value = total_price[i]
        sheet_obj.cell(row=row_start + 1, column=5).font = my_font_1
        sheet_obj.cell(row=row_start + 1, column=6).value = state[i]
        sheet_obj.cell(row=row_start + 1, column=6).font = my_font_1
        sheet_obj.cell(row=row_start + 1, column=7).value = country[i]
        sheet_obj.cell(row=row_start + 1, column=7).font = my_font_1
        sheet_obj.cell(row=row_start + 1, column=8).value = address[i]
        sheet_obj.cell(row=row_start + 1, column=8).font = my_font_1
        sheet_obj.cell(row=row_start + 1, column=9).value = phone_number[i]
        sheet_obj.cell(row=row_start + 1, column=9).font = my_font_1
        sheet_obj.cell(row=row_start + 1, column=10).value = fuel_type[i]
        sheet_obj.cell(row=row_start + 1, column=10).font = my_font_1
        sheet_obj.cell(row=row_start + 1, column=11).value = cur_gst[i]
        sheet_obj.cell(row=row_start + 1, column=11).font = my_font_1
        sheet_obj.cell(row=row_start + 1, column=12).value = created_date[i]
        sheet_obj.cell(row=row_start + 1, column=12).font = my_font_1
        sheet_obj.cell(row=row_start + 1, column=13).value = delivery_date[i]
        sheet_obj.cell(row=row_start + 1, column=13).font = my_font_1
        row_start = row_start + 1
    sheet_obj.column_dimensions["A"].width = 20.0
    sheet_obj.column_dimensions["B"].width = 20.0
    sheet_obj.column_dimensions["C"].width = 20.0
    sheet_obj.column_dimensions["D"].width = 20.0
    sheet_obj.column_dimensions["E"].width = 20.0
    sheet_obj.column_dimensions["F"].width = 20.0
    sheet_obj.column_dimensions["G"].width = 20.0
    sheet_obj.column_dimensions["H"].width = 20.0
    sheet_obj.column_dimensions["I"].width = 20.0
    sheet_obj.column_dimensions["J"].width = 20.0
    sheet_obj.column_dimensions["K"].width = 20.0
    sheet_obj.column_dimensions["L"].width = 20.0
    sheet_obj.column_dimensions["M"].width = 20.0
    pres_sheet_name = wb['Sheet']
    pres_sheet_name.title = "fuelQuote_Report_" + str(created_date_new)
    wb.save(filepath)
    file_name1 = "fuelQuote_Report_" + str(created_date_new) + ".xlsx"
    return send_file(filepath, attachment_filename=file_name1, as_attachment=True)
    return '', 204


@bp.route('/logout', methods=['GET', 'POST'])
def logout():
    if request.method == "GET":
        invalid_msg = "hidden"
        access_name = "hidden"
        access_pass = "hidden"
        session.clear()
        gc.collect()
        session['user'] = False
        return render_template("login.html", invalid_msg=invalid_msg, access_name=access_name, access_pass=access_pass)
